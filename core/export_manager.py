#!/usr/bin/env python3
from datetime import datetime

class ExportManager:
    def __init__(self, bill_manager):
        self.bill_manager = bill_manager
    
    def export_to_excel(self, filename, include_time=True, detail_mode=True, person=None, export_all_persons=False):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            
            # 确定要导出的人员
            if export_all_persons:
                export_persons = self.bill_manager.get_persons_list()
            elif person:
                export_persons = [person]
            else:
                export_persons = [self.bill_manager.get_current_person()]
            
            # 根据模式和人员数量设置工作表标题和表头
            is_multi_person_export = len(export_persons) > 1
            
            if detail_mode:
                ws.title = "账单明细"
                if is_multi_person_export:
                    # 多人导出：分类 | 人员A | 人员B | ... | 总计
                    headers = ["分类"] + export_persons + ["总计"]
                    if include_time:
                        headers.append("统计日期")
                else:
                    # 单人导出：保持原格式
                    headers = ["分类", "金额(元)", "描述"]
                    if include_time:
                        headers.append("日期")
            else:
                ws.title = "账单总账"
                if is_multi_person_export:
                    # 多人导出：分类 | 人员A | 人员B | ... | 总计
                    headers = ["分类"] + export_persons + ["总计"]
                    if include_time:
                        headers.append("统计日期")
                else:
                    # 单人导出：保持原格式
                    headers = ["分类", "总金额(元)", "描述"]
                    if include_time:
                        headers.append("统计日期")
            
            # 设置表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, size=12, color="FFFFFF")
            
            # 边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 添加数据
            row = 2
            col_count = len(headers)
            
            # 添加全局预算行
            budget_total = self.bill_manager.get_total_budget()
            
            if is_multi_person_export:
                # 多人导出：全局预算显示在总计列
                ws.cell(row=row, column=1, value="全局预算")
                # 各人员列显示为空或N/A
                for i, person in enumerate(export_persons, 2):
                    ws.cell(row=row, column=i, value="N/A")
                # 总计列显示全局预算
                ws.cell(row=row, column=len(export_persons) + 2, value=budget_total)
                if include_time:
                    ws.cell(row=row, column=len(headers), value=datetime.now().strftime('%Y-%m-%d'))
            else:
                # 单人导出：保持原格式
                ws.cell(row=row, column=1, value="全局预算")
                ws.cell(row=row, column=2, value=budget_total)
                if detail_mode:
                    ws.cell(row=row, column=3, value="全局预算总额")
                    if include_time:
                        ws.cell(row=row, column=4, value=datetime.now().strftime('%Y-%m-%d'))
                else:
                    ws.cell(row=row, column=3, value="全局预算总额")
                    if include_time:
                        ws.cell(row=row, column=4, value=datetime.now().strftime('%Y-%m-%d'))
            
            # 设置预算行样式
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            row += 1
            
            # 添加各分类数据
            if is_multi_person_export:
                # 多人导出：根据模式选择不同的导出格式
                if detail_mode:
                    # 明细模式：显示每个人员每个分类的详细账单
                    self._export_multi_person_detail_format(ws, export_persons, row, col_count, thin_border, include_time)
                else:
                    # 总账模式：显示每个人员每个分类的汇总金额
                    self._export_multi_person_summary_format(ws, export_persons, row, col_count, thin_border, include_time)
                # 多人导出完成，直接跳转到列宽调整和保存
            elif detail_mode:
                # 明细模式：显示每一笔账单
                for person_name in export_persons:
                    if person_name not in self.bill_manager.bills:
                        continue
                        
                    person_bills = self.bill_manager.bills[person_name]
                    
                    for category in self.bill_manager.categories:
                        if category == "总预算":
                            continue
                        
                        bills = person_bills.get(category, [])
                        
                        if not bills:  # 如果没有账单，在单人模式下显示该分类
                            if len(export_persons) == 1:
                                ws.cell(row=row, column=1, value=category)
                                col_idx = 2
                                ws.cell(row=row, column=col_idx, value=0.00)
                                col_idx += 1
                                ws.cell(row=row, column=col_idx, value="暂无支出")
                                col_idx += 1
                                if include_time:
                                    ws.cell(row=row, column=col_idx, value=datetime.now().strftime('%Y-%m-%d'))
                                
                                # 设置样式
                                for col in range(1, col_count + 1):
                                    cell = ws.cell(row=row, column=col)
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    cell.border = thin_border
                                row += 1
                        else:
                            for bill in bills:
                                ws.cell(row=row, column=1, value=category)
                                col_idx = 2
                                ws.cell(row=row, column=col_idx, value=bill["amount"])
                                col_idx += 1
                                ws.cell(row=row, column=col_idx, value=bill.get("description", ""))
                                col_idx += 1
                                if include_time:
                                    ws.cell(row=row, column=col_idx, value=bill["date"])
                                
                                # 设置样式
                                for col in range(1, col_count + 1):
                                    cell = ws.cell(row=row, column=col)
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    cell.border = thin_border
                                row += 1
            else:
                # 总账模式：显示各分类的汇总金额和分类备注
                if export_all_persons and len(export_persons) > 1:
                    # 导出所有人员时，按分类汇总所有人员的数据
                    for category in self.bill_manager.categories:
                        if category == "总预算":
                            continue
                        
                        category_total = self.bill_manager.get_all_persons_category_total(category)
                        description = self.bill_manager.get_category_note(category)
                        
                        ws.cell(row=row, column=1, value=category)
                        col_idx = 2
                        ws.cell(row=row, column=col_idx, value=category_total)
                        col_idx += 1
                        ws.cell(row=row, column=col_idx, value=description)
                        col_idx += 1
                        if include_time:
                            ws.cell(row=row, column=col_idx, value=datetime.now().strftime('%Y-%m-%d'))
                        
                        # 设置样式
                        for col in range(1, col_count + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border
                        row += 1
                else:
                    # 按人员分别显示各分类的汇总
                    for person_name in export_persons:
                        if person_name not in self.bill_manager.bills:
                            continue
                            
                        person_bills = self.bill_manager.bills[person_name]
                        
                        for category in self.bill_manager.categories:
                            if category == "总预算":
                                continue
                            
                            bills = person_bills.get(category, [])
                            category_total = sum(bill["amount"] for bill in bills)
                            description = self.bill_manager.get_category_note(category)
                            
                            ws.cell(row=row, column=1, value=category)
                            col_idx = 2
                            ws.cell(row=row, column=col_idx, value=category_total)
                            col_idx += 1
                            ws.cell(row=row, column=col_idx, value=description)
                            col_idx += 1
                            if include_time:
                                ws.cell(row=row, column=col_idx, value=datetime.now().strftime('%Y-%m-%d'))
                            
                            # 设置样式
                            for col in range(1, col_count + 1):
                                cell = ws.cell(row=row, column=col)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.border = thin_border
                                if category_total > 0:
                                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                            row += 1
            
            # 单人导出的汇总行（多人导出已在_export_multi_person_format中处理）
            if not is_multi_person_export:
                total_spent = self.bill_manager.get_total_spent()
                balance = self.bill_manager.get_remaining_budget()
                
                ws.cell(row=row, column=1, value="支出汇总")
                ws.cell(row=row, column=2, value=total_spent)
                
                if detail_mode:
                    ws.cell(row=row, column=3, value="所有分类支出总计")
                else:
                    ws.cell(row=row, column=3, value="所有分类支出总计")
                
                if include_time:
                    ws.cell(row=row, column=4, value=datetime.now().strftime('%Y-%m-%d'))
                
                # 设置汇总行样式
                for col in range(1, col_count + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    cell.font = Font(bold=True)
                row += 1
                
                # 添加余额行
                ws.cell(row=row, column=1, value="剩余余额")
                ws.cell(row=row, column=2, value=balance)
                
                if detail_mode:
                    ws.cell(row=row, column=3, value="预算减去支出" if balance >= 0 else "超支金额")
                else:
                    ws.cell(row=row, column=3, value="预算减去支出" if balance >= 0 else "超支金额")
                
                if include_time:
                    ws.cell(row=row, column=4, value=datetime.now().strftime('%Y-%m-%d'))
                
                # 设置余额行样式
                fill_color = "90EE90" if balance >= 0 else "FFB6C1"
                for col in range(1, col_count + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = Font(bold=True)
            
            # 自动调整列宽 - 改进版
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        # 计算中文字符，中文字符宽度按2计算
                        char_length = 0
                        for char in cell_value:
                            if '\u4e00' <= char <= '\u9fff':  # 中文字符范围
                                char_length += 2
                            else:
                                char_length += 1
                        
                        if char_length > max_length:
                            max_length = char_length
                    except:
                        pass
                
                # 设置最小和最大宽度，并增加更多缓冲空间
                min_width = 10
                max_width = 80
                adjusted_width = max(min_width, min(max_length + 8, max_width))
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            return True
            
        except Exception as e:
            return False
    
    def export_to_png(self, filename, include_time=True, detail_mode=True):
        try:
            import matplotlib.pyplot as plt

            categories = []
            amounts = []
            
            for category in self.bill_manager.categories:
                if category != "总预算":
                    total = self.bill_manager.get_category_total(category)
                    if total > 0:
                        categories.append(category)
                        amounts.append(total)
            
            if not categories:
                return False
            
            plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'PingFang SC', 'Heiti TC', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))
            fig.patch.set_facecolor('#f5f8f7')
            
            colors = plt.cm.Set2(range(len(categories)))
            
            # 饼图
            ax1.pie(amounts, labels=categories, autopct='%1.1f%%', 
                   colors=colors, textprops={'color': '#17312d'})
            title1 = f'支出分布 - {"明细模式" if detail_mode else "总账模式"}'
            ax1.set_title(title1, color='#17312d', fontsize=16, fontweight='bold')
            ax1.set_facecolor('#f5f8f7')
            
            # 柱状图
            bars = ax2.bar(categories, amounts, color=colors)
            title2 = f'各类别支出金额 - {"明细模式" if detail_mode else "总账模式"}'
            ax2.set_title(title2, color='#17312d', fontsize=16, fontweight='bold')
            ax2.set_xlabel('分类', color='#64748b')
            ax2.set_ylabel('金额 (元)', color='#64748b')
            ax2.tick_params(colors='#17312d')
            ax2.set_facecolor('#ffffff')
            ax2.grid(axis='y', color='#d9e6e2', linestyle='-', linewidth=0.8)
            ax2.spines['top'].set_visible(False)
            ax2.spines['right'].set_visible(False)
            ax2.spines['left'].set_color('#b8d5ce')
            ax2.spines['bottom'].set_color('#b8d5ce')
            
            # 添加数值标签
            for bar, amount in zip(bars, amounts):
                ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(amounts)*0.01,
                        f'¥{amount:.2f}', ha='center', va='bottom', color='#17312d')
            
            # 如果包含时间，在图表底部添加时间戳
            if include_time:
                fig.suptitle(f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 
                           color='#64748b', fontsize=12, y=0.02)
            
            plt.xticks(rotation=45)
            plt.tight_layout()
            if include_time:
                plt.subplots_adjust(bottom=0.1)  # 为时间戳留出空间
            plt.savefig(filename, facecolor='#f5f8f7', dpi=300, bbox_inches='tight')
            plt.close(fig)
            return True
            
        except Exception as e:
            return False
    
    def _export_multi_person_detail_format(self, ws, export_persons, start_row, col_count, thin_border, include_time):
        """
        多人明细模式导出：显示每个人员每个分类的具体账单明细
        格式：分类 | 人员 | 金额 | 描述 | 日期
        """
        from openpyxl.styles import Alignment, PatternFill, Font
        
        row = start_row
        
        # 重新设计表头，因为明细模式需要不同的列
        # 清除原有表头，重新设置
        for col in range(1, col_count + 1):
            ws.cell(row=1, column=col, value="")
        
        # 明细模式表头：分类 | 人员 | 金额 | 描述 | [日期]
        headers = ["分类", "人员", "金额(元)", "描述"]
        if include_time:
            headers.append("日期")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # 重置col_count
        col_count = len(headers)
        
        # 重新添加全局预算行
        row = 2
        ws.cell(row=row, column=1, value="全局预算")
        ws.cell(row=row, column=2, value="全局")
        ws.cell(row=row, column=3, value=self.bill_manager.get_total_budget())
        ws.cell(row=row, column=4, value="全局预算总额")
        if include_time:
            ws.cell(row=row, column=5, value=datetime.now().strftime('%Y-%m-%d'))
        
        # 设置预算行样式
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        row += 1
        
        # 显示每个人员的每笔账单
        for person_name in export_persons:
            if person_name not in self.bill_manager.bills:
                continue
                
            person_bills = self.bill_manager.bills[person_name]
            
            for category in self.bill_manager.categories:
                if category == "总预算":
                    continue
                
                bills = person_bills.get(category, [])
                
                if not bills:  # 如果没有账单，显示该分类为0
                    ws.cell(row=row, column=1, value=category)
                    ws.cell(row=row, column=2, value=person_name)
                    ws.cell(row=row, column=3, value=0.00)
                    ws.cell(row=row, column=4, value="暂无支出")
                    if include_time:
                        ws.cell(row=row, column=5, value=datetime.now().strftime('%Y-%m-%d'))
                    
                    # 设置样式
                    for col in range(1, col_count + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    row += 1
                else:
                    for bill in bills:
                        ws.cell(row=row, column=1, value=category)
                        ws.cell(row=row, column=2, value=person_name)
                        ws.cell(row=row, column=3, value=bill["amount"])
                        ws.cell(row=row, column=4, value=bill.get("description", ""))
                        if include_time:
                            ws.cell(row=row, column=5, value=bill["date"])
                        
                        # 设置样式
                        for col in range(1, col_count + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border
                        row += 1
    
    def _export_multi_person_summary_format(self, ws, export_persons, start_row, col_count, thin_border, include_time):
        """
        多人总账模式导出：显示每个人员每个分类的汇总金额
        格式：分类 | 人员A | 人员B | ... | 总计 | [统计日期]
        """
        from openpyxl.styles import Alignment, PatternFill, Font
        
        row = start_row
        
        # 为每个分类创建一行
        for category in self.bill_manager.categories:
            if category == "总预算":
                continue
            
            ws.cell(row=row, column=1, value=category)
            
            # 计算各人员在该分类的支出
            category_totals = []
            total_for_category = 0
            
            for i, person_name in enumerate(export_persons, 2):
                person_total = self.bill_manager.get_category_total(category, person_name)
                ws.cell(row=row, column=i, value=person_total)
                category_totals.append(person_total)
                total_for_category += person_total
            
            # 总计列
            ws.cell(row=row, column=len(export_persons) + 2, value=total_for_category)
            
            # 时间列（如果需要）
            if include_time:
                ws.cell(row=row, column=len(export_persons) + 3, value=datetime.now().strftime('%Y-%m-%d'))
            
            # 设置样式
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # 如果该分类有支出，高亮显示
                if total_for_category > 0:
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            row += 1
        
        # 添加支出汇总行
        ws.cell(row=row, column=1, value="支出汇总")
        
        # 计算各人员总支出
        for i, person_name in enumerate(export_persons, 2):
            person_total_spent = 0
            for category in self.bill_manager.categories:
                if category != "总预算":
                    person_total_spent += self.bill_manager.get_category_total(category, person_name)
            ws.cell(row=row, column=i, value=person_total_spent)
        
        # 所有人员总支出
        total_spent = self.bill_manager.get_total_spent()
        ws.cell(row=row, column=len(export_persons) + 2, value=total_spent)
        
        if include_time:
            ws.cell(row=row, column=len(export_persons) + 3, value=datetime.now().strftime('%Y-%m-%d'))
        
        # 设置汇总行样式
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
            cell.font = Font(bold=True)
        
        row += 1
        
        # 添加剩余预算行
        ws.cell(row=row, column=1, value="剩余预算")
        
        # 剩余预算显示在总计列
        remaining_budget = self.bill_manager.get_remaining_budget()
        ws.cell(row=row, column=len(export_persons) + 2, value=remaining_budget)
        
        # 其他人员列显示为N/A或空
        for i in range(2, len(export_persons) + 2):
            ws.cell(row=row, column=i, value="N/A")
        
        if include_time:
            ws.cell(row=row, column=len(export_persons) + 3, value=datetime.now().strftime('%Y-%m-%d'))
        
        # 设置剩余预算行样式
        fill_color = "90EE90" if remaining_budget >= 0 else "FFB6C1"
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = Font(bold=True)
        
        return row + 1
